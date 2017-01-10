import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet import dimensions
from openpyxl.utils import get_column_letter

home_dir = os.path.dirname(os.path.abspath(__file__))
wb = Workbook()
dest_filename = "RTAA_FR.xlsx"

ws1 = wb.active
ws1.title = "Introduction"
ws2 = wb.create_sheet(title="Framework")
ws3 = wb.create_sheet(title="Data Viewer Application")
ws4 = wb.create_sheet(title="eDoc Discovery")
sheets = [ws1, ws2, ws3, ws4]

title_fnt = Font(bold=True, size=14)
title_al = Alignment(horizontal="center", vertical="center")
title_fill = PatternFill("solid", fgColor="DDDDDD")
for ws in sheets:
    # insert the title bar for each page
    a1 = ws['A1']
    a1.value = "FUNCTIONAL REQUIREMENTS - {}".format(ws.title)
    a1.font = title_fnt
    a1.alignment = title_al
    a1.fill = title_fill
    ws.merge_cells('A1:F1')

# Write the intro summary on sheet 1
intro = ws1['A2']
intro_al = Alignment(horizontal="left", vertical="top", wrap_text=True)
intro.value = "These GIS Web Applications are to be designed to assist airport staff at the Reno Tahoe Airport.  "\
    "The applications included in this Function Requirements (FR) worksheet are: \n\n1) {} \n2) {} \n3) {}".format(ws2.title,
                                                                                                             ws3.title,
                                                                                                                   ws4.title)
intro.alignment = intro_al
ws1.merge_cells('A2:F15')
ws1.column_dimensions['A'].width = 50

# work with the FR sheets
header_fnt = Font(name='Calibri',
                  size=11,
                  bold=True)
for ws in sheets[1:]:
    # insert column names starting
    ws['A2'].value = "Req #"
    ws['B2'].value = "Requirement Summary"
    ws['C2'].value = "Completion Status"
    ws['D2'].value = "Importance Level"
    ws['E2'].value = "Technical Dependencies"
    ws['F2'].value = "Comments"

    for i in range(ws.max_column):
        ws.column_dimensions[get_column_letter(i+1)].width = 25
    for row in ws.iter_rows(min_row=2, max_col=6, max_row=2):
        for cell in row:
            cell.font = header_fnt
wb.save(filename=dest_filename)
